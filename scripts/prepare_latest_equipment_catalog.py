#!/usr/bin/env python3
"""Extract the current equipment catalogue's row-linked images and price data."""

from __future__ import annotations

import argparse
import json
import posixpath
import re
import shutil
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import openpyxl


NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "etc": "http://www.wps.cn/officeDocument/2017/etCustomData",
}
RID = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
SKU_RE = re.compile(r"^Y\d+$", re.IGNORECASE)
IMAGE_ID_RE = re.compile(r'DISPIMG\("(ID_[A-F0-9]+)"')


def relationship_map(root: ET.Element) -> dict[str, str]:
    return {
        relation.attrib["Id"]: relation.attrib["Target"]
        for relation in root.findall("rel:Relationship", NS)
    }


def archive_target(base: str, target: str) -> str:
    return posixpath.normpath(posixpath.join(base, target))


def numeric(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def row_image_targets(
    book: zipfile.ZipFile, sku_rows_by_sheet: dict[int, set[int]]
) -> dict[tuple[int, int], str]:
    cell_relationships = relationship_map(
        ET.fromstring(book.read("xl/_rels/cellimages.xml.rels")),
    )
    cell_images = {
        picture.find(".//xdr:cNvPr", NS).attrib["name"]: archive_target(
            "xl", cell_relationships[picture.find(".//a:blip", NS).attrib[RID]],
        )
        for picture in ET.fromstring(book.read("xl/cellimages.xml")).findall(
            "etc:cellImage", NS,
        )
    }
    targets: dict[tuple[int, int], str] = {}
    for sheet_number in sku_rows_by_sheet:
        sheet = ET.fromstring(book.read(f"xl/worksheets/sheet{sheet_number}.xml"))
        for cell in sheet.findall(".//x:c", NS):
            coordinate = cell.attrib.get("r", "")
            if not coordinate.startswith("C"):
                continue
            formula = cell.find("x:f", NS)
            matched = IMAGE_ID_RE.search(formula.text or "") if formula is not None else None
            if matched:
                row = int(re.search(r"\d+$", coordinate).group(0))
                targets[(sheet_number, row)] = cell_images[matched.group(1)]

        drawing_rel_path = f"xl/worksheets/_rels/sheet{sheet_number}.xml.rels"
        if drawing_rel_path not in book.namelist():
            continue
        sheet_rels = relationship_map(ET.fromstring(book.read(drawing_rel_path)))
        drawing_target = next(
            (
                target
                for target in sheet_rels.values()
                if target.startswith("../drawings/")
            ),
            None,
        )
        if not drawing_target:
            continue
        drawing_path = archive_target("xl/worksheets", drawing_target)
        drawing_rel_name = (
            f"{posixpath.dirname(drawing_path)}/_rels/{posixpath.basename(drawing_path)}.rels"
        )
        drawing_rels = relationship_map(ET.fromstring(book.read(drawing_rel_name)))
        drawing = ET.fromstring(book.read(drawing_path))
        deferred_targets: list[tuple[int, str]] = []
        for anchor in drawing.findall("xdr:twoCellAnchor", NS):
            start = anchor.find("xdr:from", NS)
            picture = anchor.find("xdr:pic", NS)
            if start is None or picture is None:
                continue
            column = start.findtext("xdr:col", namespaces=NS)
            row = start.findtext("xdr:row", namespaces=NS)
            blip = picture.find(".//a:blip", NS)
            if column not in {"1", "2"} or row is None or blip is None:
                continue
            relation_id = blip.attrib[RID]
            target = archive_target(
                posixpath.dirname(drawing_path), drawing_rels[relation_id]
            )
            row_number = int(row) + 1
            if column == "2":
                targets[(sheet_number, row_number)] = target
            else:
                deferred_targets.append((row_number, target))
        sku_rows = sku_rows_by_sheet[sheet_number]
        # WPS occasionally anchors a source image in column B. Keep it on the
        # matching product row; a B-column image on an area heading belongs to
        # the next product row.
        for row_number, target in deferred_targets:
            target_row = row_number
            if target_row not in sku_rows:
                target_row += 1
            targets.setdefault((sheet_number, target_row), target)
    return targets


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    args.output.mkdir(parents=True, exist_ok=True)
    images_dir = args.output / "images"
    if images_dir.exists():
        shutil.rmtree(images_dir)
    images_dir.mkdir()

    with zipfile.ZipFile(args.workbook) as archive:
        source = openpyxl.load_workbook(args.workbook, data_only=False)
        values = openpyxl.load_workbook(args.workbook, data_only=True)
        sku_rows_by_sheet = {
            sheet_number: {
                row
                for row in range(1, sheet.max_row + 1)
                if SKU_RE.fullmatch(clean_text(sheet.cell(row, 1).value).upper())
            }
            for sheet_number, sheet in enumerate(source.worksheets, start=1)
        }
        image_targets = row_image_targets(archive, sku_rows_by_sheet)
        products: list[dict[str, Any]] = []
        seen_image_sources: dict[str, str] = {}

        for sheet_number, sheet in enumerate(source.worksheets, start=1):
            value_sheet = values[sheet.title]
            kind = "junior" if "小童" in sheet.title else "senior"
            for row in range(1, sheet.max_row + 1):
                sku = clean_text(sheet.cell(row, 1).value).upper()
                if not SKU_RE.fullmatch(sku):
                    continue
                source_image = image_targets.get((sheet_number, row), "")
                image_file = ""
                if source_image:
                    suffix = Path(source_image).suffix.lower() or ".png"
                    image_file = f"{kind}-{sku}-{row}{suffix}"
                    destination = images_dir / image_file
                    if source_image not in seen_image_sources:
                        destination.write_bytes(archive.read(source_image))
                        seen_image_sources[source_image] = image_file
                    else:
                        (images_dir / image_file).hardlink_to(
                            images_dir / seen_image_sources[source_image]
                        )
                products.append(
                    {
                        "id": f"{kind}-{sku}",
                        "sku": sku,
                        "sheet": sheet.title,
                        "row": row,
                        "name": clean_text(sheet.cell(row, 2).value),
                        "factoryPrice": numeric(value_sheet.cell(row, 5).value),
                        "vipPrice": numeric(value_sheet.cell(row, 6).value),
                        "usdPrice": numeric(value_sheet.cell(row, 7).value),
                        "factoryRaw": clean_text(sheet.cell(row, 5).value),
                        "vipRaw": clean_text(sheet.cell(row, 6).value),
                        "usdRaw": clean_text(sheet.cell(row, 7).value),
                        "image": image_file,
                        "sourceImage": source_image,
                    }
                )

    products_by_sku: dict[str, list[dict[str, Any]]] = defaultdict(list)
    products_by_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for product in products:
        products_by_sku[product["sku"]].append(product)
        products_by_id[product["id"]].append(product)
    duplicate_skus = {
        sku: rows for sku, rows in products_by_sku.items() if len(rows) > 1
    }
    preferred_rows = {"junior-Y10590": 177}
    canonical_products = []
    for product_id in sorted(products_by_id):
        matches = products_by_id[product_id]
        preferred_row = preferred_rows.get(product_id)
        canonical_products.append(
            next(
                (
                    product
                    for product in matches
                    if preferred_row is not None and product["row"] == preferred_row
                ),
                matches[0],
            )
        )
    manifest = {
        "workbook": args.workbook.name,
        "products": products,
        "canonicalProducts": canonical_products,
        "duplicateSkus": duplicate_skus,
        "missingImages": [
            {"id": product["id"], "sheet": product["sheet"], "row": product["row"]}
            for product in products
            if not product["image"]
        ],
    }
    (args.output / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "products": len(products),
                "uniqueSkus": len({product["sku"] for product in products}),
                "duplicateSkus": len(duplicate_skus),
                "missingImages": len(manifest["missingImages"]),
                "images": len(list(images_dir.iterdir())),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
